from censys.search import CensysHosts
import json
import pandas as pd
import openpyxl
import socket
from openpyxl.styles import NamedStyle, PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl import load_workbook
import re
import ipaddress
from censys.common.exceptions import CensysUnauthorizedException


def is_valid_ip(ip):
    pattern = r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$"
    return bool(re.match(pattern, ip))


def is_valid_cidr(cidr):
    try:
        ipaddress.ip_network(cidr, strict=False)
        return True
    except ValueError:
        return False


def expand_cidr(cidr):
    network = ipaddress.ip_network(cidr, strict=False)
    return [str(ip) for ip in network.hosts()]


def check_tcp_port(ip, port):
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(2)
    result = sock.connect_ex((ip, port))
    sock.close()
    return 1 if result == 0 else 0


def check_udp_port(ip, port):
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(2)
    try:
        sock.sendto(b'', (ip, port))
        data, server = sock.recvfrom(1024)
        return 1
    except socket.timeout:
        return 0
    except socket.error:
        return 0
    finally:
        sock.close()


def check_port(ip, transport, port):
    return check_tcp_port(ip, port) if transport == "TCP" else 1


def read_ip_file(file_path):
    with open(file_path, 'r') as file:
        return [line.strip() for line in file.readlines()]


api_count = 0


def load_api_keys(file_path):
    global api_count
    api_keys = []
    api_count = 0  # Khởi tạo biến api_count trước khi sử dụng
    with open(file_path, 'r') as file:
        for line in file:
            line = line.strip()  # Loại bỏ khoảng trắng và ký tự xuống dòng từ dòng
            if line:  # Kiểm tra xem dòng có giá trị không trống
                parts = line.split(':')
                if len(parts) == 2:  # Định dạng apikey:secretkey
                    api_key, api_secret = parts
                    api_keys.append(("", api_key.strip(), api_secret.strip()))
                elif len(parts) == 3:  # Định dạng account:apikey:secretkey
                    account, api_key, api_secret = parts
                    api_keys.append(
                        (account.strip(), api_key.strip(), api_secret.strip()))
                else:
                    print(f"Ignoring invalid line: {line}")
                api_count += 1  # Tăng giá trị của api_count
    return api_keys


def get_censys_info(api_keys, ip_list):
    api_fall_count = 0
    extracted_data = {}
    line_of_ip = 1
    for item in ip_list:
        if is_valid_ip(item):
            ips_to_scan = [item]
        elif is_valid_cidr(item):
            ips_to_scan = expand_cidr(item)
        else:
            print("Detected a invalid IP or CIDR. See more detail in log_err.log")
            with open("log_err.log", 'a', encoding='utf-8') as file:
                file.write(
                    f"row {line_of_ip}:{item} -> Not a valid IP or CIDR\n")
            line_of_ip += 1
            continue

        for ip in ips_to_scan:
            line_of_ip += 1
            for account, api_id, api_secret in api_keys:
                if account == "":
                    account = api_id
                censys = CensysHosts(api_id=api_id, api_secret=api_secret)
                try:
                    print(f"Fetching censys by {account} for {ip}.....")
                    data = censys.view(ip)
                    domains = data.get("dns", {}).get("names", [])
                    domain = ', '.join(domains)
                    operating_system = data.get(
                        "operating_system", {}).get("product")
                    extracted_services = []

                    for service in data.get("services", []):
                        _decoded = service.get("_decoded", "")
                        port = service.get("port", "")
                        transport_protocol = service.get(
                            "transport_protocol", "").upper()
                        observed_at = service.get("observed_at", "")
                        source_ip = service.get("source_ip", "")
                        if check_port(ip, transport_protocol, port):
                            if transport_protocol == "TCP":
                                softwares = []
                                for software in service.get("software", []):
                                    if software.get("product") != operating_system:
                                        version = "_" + \
                                            software.get("version", "") if software.get(
                                                "version") else ""
                                        product = f"{software.get('product', '')}{version}"
                                        softwares.append(product)

                                extracted_service = {
                                    "port_protocol": f"{port}/{transport_protocol.upper()}",
                                    "service": _decoded,
                                    "product": softwares,
                                    "observed_at": observed_at,
                                    "source_ip": source_ip
                                }
                            else:
                                extracted_service = {
                                    "port_protocol": f"{port}/{transport_protocol.upper()}",
                                    "service": _decoded,
                                    "product": [],
                                    "observed_at": observed_at,
                                    "source_ip": source_ip
                                }
                            extracted_services.append(extracted_service)
                        else:
                            print(
                                "Detected a unused port. See more details in log_err.log")
                            with open("log_err.log", 'a', encoding='utf-8') as file:
                                file.write(
                                    f"row {line_of_ip}:ip {ip}:port {port} -> closed port\n")
                            continue

                    extracted_data[ip] = {
                        "domain": domain,
                        "services": extracted_services
                    }
                    break
                except CensysUnauthorizedException as e:
                    api_fall_count = api_fall_count + 1
                    print(f"API key error. See more details in log_err.log")
                    print(f"trying with next API of {account} ...")
                    with open("log_err.log", 'a', encoding='utf-8') as file:
                        file.write(f"API key of {account} error: {e}\n")
                    if (api_fall_count == api_count):
                        print(
                            "no API available, please check the censys accouts limit or check your connection. See more details in log_err.log\n")
                        with open("log_err.log", 'a', encoding='utf-8') as file:
                            file.write(
                                f"no API available, please check the censys accouts limit or check your connection.\n")
                        exit()
                    continue
                except Exception as e:
                    extracted_data[ip] = {"error": str(e)}
                    break

    return extracted_data


def json_to_excel(json_data, output_file):
    data_dict = json.loads(json_data)
    dfs = []

    for ip, ip_data in data_dict.items():
        domain = ip_data.get('domain', '')
        services = ip_data.get('services', [])
        for service in services:
            port_protocol = service.get('port_protocol', '')
            product = ', '.join(service.get('product', []))
            observed_at = service.get('observed_at', '')
            source_ip = service.get('source_ip', '')
            df = pd.DataFrame({
                'IP': [ip],
                'Domain': [domain],
                'Cổng/Giao Thức': [port_protocol],
                'Dịch vụ': [service.get('service', '')],
                'Product': [product],
                'Ghi Chú': [''],
                'Khuyến Nghị': ["Kiểm tra xem port có sử dụng không, nếu không sử dụng thì restrict IP được truy cập vào"],
                "Lần cuối kiểm tra": [observed_at],
                'source_ip': [source_ip]
            })
            dfs.append(df)

    if dfs:
        final_df = pd.concat(dfs, ignore_index=True)
        final_df.to_excel(output_file, index=False)
    else:
        print("No data to write into Excel file.")
        with open("log_err.log", 'a') as file:
            file.write(
                "No data to write into Excel file. May be the API call is wrong format.\n")
        exit()


def apply_table_style(ws):
    table_style = NamedStyle(name="Table")
    table_style.alignment = Alignment(horizontal='center', vertical='center')
    fill = PatternFill(start_color="F2F2F2",
                       end_color="F2F2F2", fill_type="solid")
    table_style.fill = fill


def apply_conditional_formatting(ws):
    rule = Rule(type="expression", dxf=DifferentialStyle(fill=PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid")))
    rule.formula = ["TRUE"]
    ws.conditional_formatting.add("A1:I1", rule)


def find_last_row(ws, column):
    for row in range(ws.max_row, 0, -1):
        cell_value = ws.cell(row=row, column=column).value
        if cell_value:
            return row
    return 1


def apply_border_to_header(ws):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    last_row_tmp = find_last_row(ws, 1)
    for col in range(1, 10):
        last_row = find_last_row(ws, col)
        if last_row < last_row_tmp:
            last_row = last_row_tmp
    for col in range(1, 10):
        for row in range(1, last_row + 1):
            ws.cell(row=row, column=col).border = thin_border


def autofit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column = column[0].column_letter
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


def format_excel_sheet(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    apply_table_style(ws)
    apply_conditional_formatting(ws)
    autofit_columns(ws)
    apply_border_to_header(ws)
    wb.save(file_path)


def merge(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    current_value_col1 = None
    current_value_col2 = None
    current_value_col3 = None
    start_row_col1 = None
    start_row_col2 = None
    start_row_col3 = None
    end_row_col1 = None
    end_row_col2 = None
    end_row_col3 = None

    for row in range(1, ws.max_row + 1):
        cell_value_col1 = ws.cell(row=row, column=1).value
        cell_value_col2 = ws.cell(row=row, column=2).value
        cell_value_col3 = ws.cell(row=row, column=3).value

        if cell_value_col1 == current_value_col1:
            end_row_col1 = row
        else:
            if start_row_col1 is not None and end_row_col1 is not None:
                merge_cells(ws, start_row_col1, end_row_col1, 1)
                align_cells(ws, start_row_col1, end_row_col1, 1)
            current_value_col1 = cell_value_col1
            start_row_col1 = row
            end_row_col1 = row

        if cell_value_col2 == current_value_col2:
            end_row_col2 = row
        else:
            if start_row_col2 is not None and end_row_col2 is not None:
                merge_cells(ws, start_row_col2, end_row_col2, 2)
                align_cells(ws, start_row_col2, end_row_col2, 2)
            current_value_col2 = cell_value_col2
            start_row_col2 = row
            end_row_col2 = row

        if cell_value_col3 == current_value_col3:
            end_row_col3 = row
        else:
            if start_row_col3 is not None and end_row_col3 is not None:
                merge_cells(ws, start_row_col3, end_row_col3, 3)
                merge_cells(ws, start_row_col3, end_row_col3, 5)
                merge_cells(ws, start_row_col3, end_row_col3, 6)
                merge_values(ws, start_row_col3, end_row_col3, 4)
                merge_values(ws, start_row_col3, end_row_col3, 7)
                merge_values(ws, start_row_col3, end_row_col3, 8)
                merge_values(ws, start_row_col3, end_row_col3, 9)

                align_cells(ws, start_row_col3, end_row_col3, 3)
                align_cells(ws, start_row_col3, end_row_col3, 4)
                align_cells(ws, start_row_col3, end_row_col3, 5)
                align_cells(ws, start_row_col3, end_row_col3, 6)
                align_cells(ws, start_row_col3, end_row_col3, 7)
                align_cells(ws, start_row_col3, end_row_col3, 8)
                align_cells(ws, start_row_col3, end_row_col3, 9)
            current_value_col3 = cell_value_col3
            start_row_col3 = row
            end_row_col3 = row

    if start_row_col1 is not None and end_row_col1 is not None:
        merge_cells(ws, start_row_col1, end_row_col1, 1)
        align_cells(ws, start_row_col1, end_row_col1, 1)

    if start_row_col2 is not None and end_row_col2 is not None:
        merge_cells(ws, start_row_col2, end_row_col2, 2)
        align_cells(ws, start_row_col2, end_row_col2, 2)

    if start_row_col3 is not None and end_row_col3 is not None:
        merge_cells(ws, start_row_col3, end_row_col3, 3)
        merge_cells(ws, start_row_col3, end_row_col3, 5)
        merge_cells(ws, start_row_col3, end_row_col3, 6)
        merge_values(ws, start_row_col3, end_row_col3, 4)
        merge_values(ws, start_row_col3, end_row_col3, 7)
        merge_values(ws, start_row_col3, end_row_col3, 8)
        merge_values(ws, start_row_col3, end_row_col3, 9)

        align_cells(ws, start_row_col3, end_row_col3, 3)
        align_cells(ws, start_row_col3, end_row_col3, 4)
        align_cells(ws, start_row_col3, end_row_col3, 5)
        align_cells(ws, start_row_col3, end_row_col3, 6)
        align_cells(ws, start_row_col3, end_row_col3, 7)
        align_cells(ws, start_row_col3, end_row_col3, 8)
        align_cells(ws, start_row_col3, end_row_col3, 9)

    wb.save(output_file)


def merge_values(ws, start_row, end_row, column):
    merged_value = ",".join(
        [ws.cell(row=row, column=column).value for row in range(start_row, end_row + 1)])
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=column).value = merged_value
    merge_cells(ws, start_row, end_row, column)


def merge_cells(ws, start_row, end_row, column):
    ws.merge_cells(start_row=start_row, end_row=end_row,
                   start_column=column, end_column=column)


def align_cells(ws, start_row, end_row, column):
    alignment = Alignment(horizontal='center', vertical='center')
    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=column).alignment = alignment


def main():
    ip_file = input("Input IP pass path: ")
    print("---It had 2 kind of apikey_secretkey format in API_KEYS_FILE (account:apikey:secretkey and apikey:secretkey)---")
    api_keys_file = input("Input API_KEY:SECRET_KEY path: ")

    ip_list = read_ip_file(ip_file)
    api_keys = load_api_keys(api_keys_file)

    print("Fetching data from Censys...")
    censys_info = get_censys_info(api_keys, ip_list)

    censys_info_json = json.dumps(censys_info, indent=4)
    json_to_excel(censys_info_json, 'output.xlsx')
    merge("output.xlsx", "output.xlsx")
    format_excel_sheet("output.xlsx")
    print("Process completed successfully. The Results file will be Saved into output.xlsx")


if __name__ == "__main__":
    main()
